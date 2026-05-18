"""Excel export for the warehouse registry — both lots and movements.

Two report kinds are supported:

* ``lots``      — current state of every batch (one row per Lot).
* ``movements`` — every InventoryMovement (receipts, transfers, issues, etc).

The caller chooses which columns to include and which filters to apply. The
service mirrors the same access-control and warehouse-scope rules as the
JSON endpoints so users cannot bypass scope by switching format.
"""
from __future__ import annotations

import io
from datetime import date, datetime
from typing import Iterable, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import String, cast, func, literal, or_
from sqlalchemy.orm import Session

from app.api.deps import CurrentUser
from app.models.inventory import InventoryMovement, Lot
from app.models.master_data import Location, Manufacturer, Material, Supplier, Warehouse


# Column registry: id -> (header, getter). Adding a new column == one line.
LOT_COLUMNS: dict[str, tuple[str, str]] = {
    "internal_lot": ("Внутр. серия", "internal_lot"),
    "supplier_lot": ("Серия поставщика", "supplier_lot"),
    "material_code": ("Код материала", "material_code"),
    "material_name": ("Наименование", "material_name"),
    "supplier_name": ("Поставщик", "supplier_name"),
    "manufacturer_name": ("Производитель", "manufacturer_name"),
    "warehouse_type": ("Склад", "warehouse_type"),
    "location_code": ("Зона", "location_code"),
    "rack_no": ("Стеллаж", "rack_no"),
    "sector_no": ("Сектор", "sector_no"),
    "tier_no": ("Ярус", "tier_no"),
    "place_no": ("Место", "place_no"),
    "pallet_no": ("Поддон", "pallet_no"),
    "quantity": ("Остаток", "quantity"),
    "initial_quantity": ("Начальное кол-во", "initial_quantity"),
    "unit": ("Ед.", "unit"),
    "quality_status": ("Статус ОКК", "quality_status"),
    "production_date": ("Дата производства", "production_date"),
    "expiry_date": ("Срок годности", "expiry_date"),
    "incoming_control_notified_at": ("Дата прихода", "incoming_control_notified_at"),
}


MOVEMENT_COLUMNS: dict[str, tuple[str, str]] = {
    "created_at": ("Дата/время", "created_at"),
    "movement_type": ("Тип", "movement_type"),
    "document_type": ("Документ", "document_type"),
    "internal_lot": ("Внутр. серия", "internal_lot"),
    "supplier_lot": ("Серия поставщика", "supplier_lot"),
    "material_code": ("Код материала", "material_code"),
    "material_name": ("Наименование", "material_name"),
    "quantity_delta": ("Δ Количество", "quantity_delta"),
    "quantity_after": ("Остаток после", "quantity_after"),
    "unit": ("Ед.", "unit"),
    "reason": ("Основание", "reason"),
    "workstation_id": ("АРМ", "workstation_id"),
}


def _lots_query(
    db: Session,
    user: CurrentUser,
    *,
    material: str | None,
    quality_status: str | None,
    location: str | None,
    manufacturer: str | None,
    internal_lot: str | None,
    supplier_lot: str | None,
    date_from: date | None,
    date_to: date | None,
    date_type: str,
):
    q = (
        db.query(
            Lot.internal_lot.label("internal_lot"),
            func.coalesce(Lot.supplier_lot, literal("-")).label("supplier_lot"),
            Material.code.label("material_code"),
            func.coalesce(func.nullif(Material.name, ""), Material.code).label("material_name"),
            func.coalesce(Supplier.name, "-").label("supplier_name"),
            Manufacturer.name.label("manufacturer_name"),
            Warehouse.warehouse_type.label("warehouse_type"),
            Location.code.label("location_code"),
            Lot.rack_no,
            Lot.sector_no,
            Lot.tier_no,
            Lot.place_no,
            Lot.pallet_no,
            Lot.quantity,
            Lot.initial_quantity,
            Lot.unit,
            Lot.quality_status,
            Lot.production_date,
            Lot.expiry_date,
            Lot.incoming_control_notified_at,
        )
        .join(Material, Material.id == Lot.material_id)
        .outerjoin(Supplier, Supplier.id == Lot.supplier_id)
        .join(Manufacturer, Manufacturer.id == Lot.manufacturer_id)
        .join(Warehouse, Warehouse.id == Lot.warehouse_id)
        .join(Location, Location.id == Lot.location_id)
        .order_by(Lot.created_at.desc())
    )
    if user.warehouse_scope:
        q = q.filter(Warehouse.warehouse_type == user.warehouse_scope)
    if material:
        like = f"%{material.strip().lower()}%"
        q = q.filter(or_(func.lower(Material.code).like(like), func.lower(Material.name).like(like)))
    if quality_status:
        q = q.filter(Lot.quality_status == quality_status)
    if location:
        q = q.filter(func.lower(Location.code).like(f"%{location.strip().lower()}%"))
    if manufacturer:
        q = q.filter(func.lower(Manufacturer.name).like(f"%{manufacturer.strip().lower()}%"))
    if internal_lot:
        q = q.filter(func.lower(func.coalesce(Lot.supplier_lot, Lot.internal_lot)).like(f"%{internal_lot.strip().lower()}%"))
    if supplier_lot:
        q = q.filter(func.lower(func.coalesce(Lot.supplier_lot, Lot.internal_lot)).like(f"%{supplier_lot.strip().lower()}%"))
    date_field = Lot.expiry_date if date_type == "expiry" else func.date(Lot.incoming_control_notified_at)
    if date_from:
        q = q.filter(date_field >= date_from)
    if date_to:
        q = q.filter(date_field <= date_to)
    return q


def _movements_query(
    db: Session,
    user: CurrentUser,
    *,
    material: str | None,
    internal_lot: str | None,
    supplier_lot: str | None,
    document: str | None,
    movement_type: str | None,
    date_from: date | None,
    date_to: date | None,
):
    q = (
        db.query(
            InventoryMovement.created_at,
            InventoryMovement.movement_type,
            InventoryMovement.document_type,
            Lot.internal_lot.label("internal_lot"),
            func.coalesce(Lot.supplier_lot, literal("-")).label("supplier_lot"),
            Material.code.label("material_code"),
            func.coalesce(func.nullif(Material.name, ""), Material.code).label("material_name"),
            InventoryMovement.quantity_delta,
            InventoryMovement.quantity_after,
            InventoryMovement.unit,
            InventoryMovement.reason,
            InventoryMovement.workstation_id,
        )
        .join(Lot, Lot.id == InventoryMovement.lot_id)
        .join(Material, Material.id == Lot.material_id)
        .join(Warehouse, Warehouse.id == Lot.warehouse_id)
        .order_by(InventoryMovement.created_at.desc())
    )
    if user.warehouse_scope:
        q = q.filter(Warehouse.warehouse_type == user.warehouse_scope)
    if date_from:
        q = q.filter(func.date(InventoryMovement.created_at) >= date_from)
    if date_to:
        q = q.filter(func.date(InventoryMovement.created_at) <= date_to)
    if material:
        like = f"%{material.strip().lower()}%"
        q = q.filter(or_(func.lower(Material.code).like(like), func.lower(Material.name).like(like)))
    if internal_lot:
        q = q.filter(func.lower(func.coalesce(Lot.supplier_lot, Lot.internal_lot)).like(f"%{internal_lot.strip().lower()}%"))
    if supplier_lot:
        q = q.filter(func.lower(func.coalesce(Lot.supplier_lot, Lot.internal_lot)).like(f"%{supplier_lot.strip().lower()}%"))
    if document:
        like = f"%{document.strip().lower()}%"
        q = q.filter(
            or_(
                func.lower(InventoryMovement.document_type).like(like),
                func.lower(cast(InventoryMovement.document_id, String)).like(like),
            )
        )
    if movement_type:
        q = q.filter(InventoryMovement.movement_type == movement_type)
    return q


def _format_cell(value):
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    return value


def _build_workbook(title: str, headers: Sequence[str], rows: Iterable[Sequence]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]  # Excel limit
    header_fill = PatternFill("solid", fgColor="0F172A")
    header_font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
    align = Alignment(horizontal="left", vertical="center", wrap_text=False)
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row_index, row in enumerate(rows, start=2):
        for col_index, value in enumerate(row, start=1):
            cell = ws.cell(row=row_index, column=col_index, value=_format_cell(value))
            cell.font = Font(name="Calibri", size=10)
            cell.alignment = align
    # Autofit (cheap approximation — openpyxl has no native autofit)
    for col_index, header in enumerate(headers, start=1):
        max_len = len(str(header))
        letter = get_column_letter(col_index)
        for row in ws.iter_rows(min_col=col_index, max_col=col_index, min_row=2, values_only=True):
            v = row[0]
            if v is not None:
                length = len(str(v))
                if length > max_len:
                    max_len = length
        ws.column_dimensions[letter].width = min(max_len + 2, 50)
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_lots_xlsx(
    db: Session,
    user: CurrentUser,
    *,
    columns: list[str],
    material: str | None = None,
    quality_status: str | None = None,
    location: str | None = None,
    manufacturer: str | None = None,
    internal_lot: str | None = None,
    supplier_lot: str | None = None,
    date_from: date | None = None,
    date_to: date | None = None,
    date_type: str = "arrival",
) -> bytes:
    selected = [c for c in columns if c in LOT_COLUMNS] or list(LOT_COLUMNS.keys())
    headers = [LOT_COLUMNS[c][0] for c in selected]
    q = _lots_query(
        db,
        user,
        material=material,
        quality_status=quality_status,
        location=location,
        manufacturer=manufacturer,
        internal_lot=internal_lot,
        supplier_lot=supplier_lot,
        date_from=date_from,
        date_to=date_to,
        date_type=date_type,
    )
    rows_raw = q.all()
    rows = [[getattr(row, LOT_COLUMNS[c][1]) for c in selected] for row in rows_raw]
    return _build_workbook("Партии", headers, rows)


def export_movements_xlsx(
    db: Session,
    user: CurrentUser,
    *,
    columns: list[str],
    material: str | None = None,
    internal_lot: str | None = None,
    supplier_lot: str | None = None,
    document: str | None = None,
    movement_type: str | None = None,
    date_from: date | None = None,
    date_to: date | None = None,
) -> bytes:
    selected = [c for c in columns if c in MOVEMENT_COLUMNS] or list(MOVEMENT_COLUMNS.keys())
    headers = [MOVEMENT_COLUMNS[c][0] for c in selected]
    q = _movements_query(
        db,
        user,
        material=material,
        internal_lot=internal_lot,
        supplier_lot=supplier_lot,
        document=document,
        movement_type=movement_type,
        date_from=date_from,
        date_to=date_to,
    )
    rows_raw = q.all()
    rows = [[getattr(row, MOVEMENT_COLUMNS[c][1]) for c in selected] for row in rows_raw]
    return _build_workbook("Движения", headers, rows)
